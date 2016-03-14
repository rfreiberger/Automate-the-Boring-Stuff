#! python3

# sendDuesReminders.py - Sends emails based on payment status in spreadsheet. 

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status.

wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.get_sheet_by_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

# TODO: Check each member's payment status.

# Check each member's payment status.

unpaidMembers = {}

for r in range(2, sheet.get_highest_row() +1):
	payment = sheet.cell(row=r, column=lastCol).value
	if payment != 'paid':
		name = sheet.cell(row=r, column=1).value
		email = sheet.cell(row=r, column=2).value
		unpaidMembers[name] = email


# TODO: Log in to email account.

# TODO: Send out reminder emails.

